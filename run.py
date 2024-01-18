import openpyxl
from openpyxl import load_workbook

def access_file():
    """
    The function requests the user to provide name and path of the Excel spreadsheet file containing up-to-date trading data of the stock.
    """    
    print("Enter path to the folder with your data.\nCopy-paste the path from your explorer to avoid typos:")
    folder_path = input()
    print("What is the name of the share price/traded volume file?")
    file_name = input()
    file_path = folder_path +  "\\" + file_name + ".xlsx"
    #print(file_path)
    file = load_workbook(filename=file_path)
    sheet = file['Sheet1']
    #print(file.sheetnames)
    #for row in sheet.iter_rows(values_only=True):
    #    print(row)
    #    print(type(row))
    #print(type(sheet))    

def main():
    """
    Run all program functions
    """
    access_file()


print("Welcome to the 'Equity Stock Pulse Check' project.\nThis little tool will calculate Return on Investment (ROI) for your share stock based on the date of acquisition and current share price.\nIt will also suggest cell/buy strategy deduced from five-day share price and traded volume extrapolation.\nDISCLAIMER: The algorithm does not take into account unpredictable events, such as breaking news or paiment of dividends.\nTHE RESULTS CANNOT BE TREATED AS A LEGAL FINANCIAL ADVICE.\n")

main()